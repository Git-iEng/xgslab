from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.urls import reverse
from django.conf import settings
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.core.mail import get_connection, EmailMultiAlternatives
from django.utils import timezone
from django.contrib import messages

from pathlib import Path
from datetime import datetime
from threading import Thread
import os, re
import pandas as pd
import phonenumbers
import pycountry
from phonenumbers import PhoneNumberFormat
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .forms import ContactForm
from .utils_excel import append_submission_xlsx
from .utils_contact import normalize_phone_and_country, country_name_from_alpha2
from django.http import JsonResponse
import pycountry, phonenumbers


# ---------- Validation patterns ----------
NAME_RE  = re.compile(r"^[A-Za-z\s'.-]{2,}$")
PHONE_RE = re.compile(r"^\+?\d[\d\s\-()]{6,}$")

# ---------- Excel paths ----------
# EXCEL_DIR  = os.path.join(settings.BASE_DIR, "data")
# EXCEL_PATH = os.path.join(EXCEL_DIR, "carl_demo_requests.xlsx")


# def _append_to_excel(row):
#     """Create/append to the Request Demo workbook."""
#     os.makedirs(EXCEL_DIR, exist_ok=True)
#     if os.path.exists(EXCEL_PATH):
#         wb = load_workbook(EXCEL_PATH)
#         ws = wb.active
#     else:
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Requests"
#         headers = [
#             "Timestamp", "Full Name", "Company", "Email",
#             "Country", "Dial Code", "Phone", "Address",
#             "Message", "Source IP",
#         ]
#         ws.append(headers)
#         for i in range(1, len(headers) + 1):
#             ws.column_dimensions[get_column_letter(i)].width = 24
#     ws.append(row)
#     wb.save(EXCEL_PATH)


# ---------- Email helpers ----------
def _send_email(subject: str, text_body: str, html_body: str | None, recipients: list[str] | None):
    """Low-level sender used by async wrappers."""
    try:
        if not recipients:
            # last-resort fallback
            fallback = getattr(settings, "EMAIL_HOST_USER", None) or getattr(settings, "DEFAULT_FROM_EMAIL", None)
            recipients = [fallback] if fallback else []

        if not recipients:
            print("EMAIL WARNING: no recipients configured")
            return

        conn = get_connection(timeout=getattr(settings, "EMAIL_TIMEOUT", 15))
        msg = EmailMultiAlternatives(
            subject=subject,
            body=text_body,
            from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None) or getattr(settings, "EMAIL_HOST_USER", None),
            to=recipients,
            connection=conn,
        )
        if html_body:
            msg.attach_alternative(html_body, "text/html")
        msg.send(fail_silently=False)
    except Exception as e:
        print("EMAIL ERROR:", repr(e))


def _send_demo_email_async(subject: str, text_body: str, html_body: str | None = None):
    """Fire-and-forget email for Request Demo."""
    recipients = getattr(settings, "DEMO_RECIPIENTS", None) or getattr(settings, "CONTACT_RECIPIENTS", None)
    Thread(target=_send_email, args=(subject, text_body, html_body, recipients), daemon=True).start()


def _send_contact_email_async(subject: str, text_body: str, html_body: str | None = None):
    """Fire-and-forget email for Contact form."""
    recipients = getattr(settings, "CONTACT_RECIPIENTS", None)
    Thread(target=_send_email, args=(subject, text_body, html_body, recipients), daemon=True).start()


# ---------- Views ----------
def request_demo_view(request):
    if request.method != "POST":
        return redirect("/")

    # Pull fields
    full_name = request.POST.get("full_name", "").strip()
    company   = request.POST.get("company", "").strip()
    email     = request.POST.get("email", "").strip()
    phone     = request.POST.get("phone", "").strip()
    country   = request.POST.get("country", "").strip()  # "IN|+91"
    address   = request.POST.get("address", "").strip()
    message   = request.POST.get("message", "").strip()

    # Validate
    errors = {}
    if not NAME_RE.match(full_name):
        errors["full_name"] = "Please enter a valid full name (letters only)."
    if not company:
        errors["company"] = "Company is required."
    try:
        validate_email(email)
    except ValidationError:
        errors["email"] = "Enter a valid email."
    if not PHONE_RE.match(phone):
        errors["phone"] = "Enter a valid phone number."
    if not country:
        errors["country"] = "Select a country."

    if errors:
        # Raise toasts on next page load
        for msg in errors.values():
            messages.error(request, msg)
        # Go back to the page that opened the modal (so your JS toast can show)
        return redirect(request.META.get("HTTP_REFERER", "/"))

    # Split "IN|+91"
    country_code, dial = (country.split("|", 1) + [""])[:2]

    # Excel append
    # _append_to_excel([
    #     timezone.now().strftime("%Y-%m-%d %H:%M:%S %Z") or timezone.now().strftime("%Y-%m-%d %H:%M:%S"),
    #     full_name, company, email, country_code, dial, phone, address, message,
    #     request.META.get("REMOTE_ADDR", ""),
    # ])

    # Build email
    ts = timezone.now().strftime("%Y-%m-%d %H:%M:%S %Z")
    subject = "New CARL Demo Request"
    text_body = (
        "A new CARL demo request was submitted.\n\n"
        f"Submitted: {ts}\n"
        f"IP: {request.META.get('REMOTE_ADDR','')}\n\n"
        f"Full name: {full_name}\n"
        f"Company: {company}\n"
        f"Email: {email}\n"
        f"Phone: {phone}\n"
        f"Country: {country_code} {dial}\n"
        f"Address: {address}\n\n"
        "Message:\n"
        f"{message or '(none)'}\n"
    )
    html_body = f"""
        <h2 style="margin:0 0 8px">New CARL Demo Request</h2>
        <p style="margin:0 0 12px;color:#334">Submitted {ts} from {request.META.get('REMOTE_ADDR','')}</p>
        <table cellpadding="6" cellspacing="0" style="border-collapse:collapse;background:#f9fbfc">
          <tr><td><b>Full name</b></td><td>{full_name}</td></tr>
          <tr><td><b>Company</b></td><td>{company}</td></tr>
          <tr><td><b>Email</b></td><td>{email}</td></tr>
          <tr><td><b>Phone</b></td><td>{phone}</td></tr>
          <tr><td><b>Country</b></td><td>{country_code} {dial}</td></tr>
          <tr><td><b>Address</b></td><td>{address}</td></tr>
        </table>
        <p style="margin:12px 0 4px"><b>Message</b></p>
        <pre style="white-space:pre-wrap;font-family:system-ui,Segoe UI,Arial,sans-serif">{message or '(none)'}</pre>
    """

    _send_demo_email_async(subject, text_body, html_body)

    # Success -> thanks page
    return redirect(reverse("cmmsApp:contact_thanks"))


def home(request):
    return render(request, "index.html")


def request_demo(request):
    return render(request, "request_demo_modal.html")


def gsa(request):    return render(request, "gsa.html")


def contact(request):     return render(request, "contact.html")

def about(request):       return render(request, "about.html")

def product(request):     return render(request, "product.html")
def services(request):     return render(request, "services.html")
def neplan_asset_management(request):     return render(request, "neplan-asset-management.html")
def nets(request):     return render(request, "nets.html")

def gsafd(request):     return render(request, "gsafd.html")
def xgsafd(request):     return render(request, "xgsafd.html")
def sheilda(request):     return render(request, "sheilda.html")
def xgsatd(request):     return render(request, "xgsatd.html")
def sheild(request):     return render(request, "sheild.html")


def contact(request):     return render(request, "neplan-contact.html")
def contact_section(request):
    form = ContactForm(request.POST or None)

    if request.method == "POST" and not form.is_valid():
        messages.error(request, "Please correct the highlighted fields and resubmit.")

    if request.method == "POST" and form.is_valid():
        cd = form.cleaned_data

        # Normalize phone & resolve country name
        e164_phone, resolved_alpha2, resolved_country_name = normalize_phone_and_country(
            cd.get("phone", ""), cd.get("country", "")
        )

        # Append to Excel
        # xlsx_path = Path(
        #     getattr(settings, "CONTACT_SUBMISSIONS_XLSX", Path(settings.BASE_DIR) / "contact_submissions.xlsx")
        # )
        # append_submission_xlsx(
        #     xlsx_path,
        #     [
        #         datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        #         cd["first_name"],
        #         cd.get("last_name", ""),
        #         cd.get("company", ""),
        #         cd["email"],
        #         resolved_alpha2,
        #         resolved_country_name,
        #         e164_phone or cd.get("phone", ""),
        #         cd.get("message", ""),
        #     ],
        # )

        # Email body
        subject = "New website contact submission for CARL Software"
        text_body = "\n".join(
            [
                "New contact submission for CARL Software:",
                f"Name: {cd['first_name']} {cd.get('last_name','')}".strip(),
                f"Company: {cd.get('company','')}",
                f"Email: {cd['email']}",
                f"Country: {resolved_country_name or country_name_from_alpha2(resolved_alpha2) or cd.get('country','')}",
                f"Phone: {e164_phone or cd.get('phone','')}",
                "",
                "Message:",
                cd.get("message", ""),
            ]
        )

        _send_contact_email_async(subject, text_body, None)

        return redirect(reverse("cmmsApp:contact_thanks"))

    return render(request, "contact_section.html", {"form": form, "sent": request.GET.get("sent")})


# ---------- NEW: helper (not a view) ----------
def _dial_code_from_alpha2(alpha2: str) -> str:
    """Return '+<code>' from a country alpha2 code."""
    if not alpha2:
        return ""
    try:
        cc = phonenumbers.country_code_for_region(alpha2.upper())
        return f"+{cc}" if cc else ""
    except Exception:
        return ""


# ---------- NEW: JSON helper endpoint ----------
def phone_info(request):
    """
    Optional helper called by the form JS to keep Country <-> Phone in sync.
    Accepts ?phone=+.. OR ?country=Name/Alpha2
    Returns: e164 phone, country (full name), alpha2, dial_code, example
    """
    phone = (request.GET.get("phone") or "").strip()
    country = (request.GET.get("country") or "").strip()

    e164, resolved_alpha2, resolved_country_name = normalize_phone_and_country(phone, country)
    dial = _dial_code_from_alpha2(resolved_alpha2)

    # simple example for UI: prefill with a dial code if user typed only country
    example = ""
    if dial and phone and not phone.startswith("+"):
        example = f"{dial} 4xxxxxxxx"
    elif dial and not phone:
        example = f"{dial} 4xxxxxxxx"

    return JsonResponse({
        "e164": e164,
        "country": resolved_country_name,
        "alpha2": resolved_alpha2,
        "dial_code": dial,
        "example": example
    })


# ---------- NEW: consulting/contact form submit ----------
def contact_block_submit(request):
    """
    Handles the 'Get Free Consulting' form shown in the new block.
    - Normalizes Country <-> Phone
    - Appends a row to CONTACT_SUBMISSIONS_XLSX
    - Sends email to CONTACT_RECIPIENTS
    """
    if request.method != "POST":
        return redirect(request.META.get("HTTP_REFERER", "/"))

    name    = (request.POST.get("name")    or "").strip()
    email   = (request.POST.get("email")   or "").strip()
    phone   = (request.POST.get("phone")   or "").strip()
    country = (request.POST.get("country") or "").strip()
    service = (request.POST.get("service") or "").strip()
    message = (request.POST.get("message") or "").strip()

    # --- Basic validation (lightweight) ---
    errors = []
    if not re.match(r"^[A-Za-z\s'.-]{2,}$", name):
        errors.append("Please enter a valid name.")
    try:
        validate_email(email)
    except ValidationError:
        errors.append("Enter a valid email address.")
    if not re.match(r"^\+?\d[\d\s\-()]{6,}$", phone):
        errors.append("Enter a valid phone number.")
    if not country and not phone.startswith("+"):
        # If there's no +code in phone, we do need a country hint
        errors.append("Please enter your country.")

    if errors:
        for e in errors:
            messages.error(request, e)
        return redirect(request.META.get("HTTP_REFERER", "/"))

    # --- Normalize country/phone ---
    e164_phone, alpha2, country_name = normalize_phone_and_country(phone, country)
    dial_code = _dial_code_from_alpha2(alpha2)

    # --- Append to Excel ---
    xlsx_path = Path(getattr(settings, "CONTACT_SUBMISSIONS_XLSX",
                             Path(settings.BASE_DIR) / "contact_submissions.xlsx"))
    # Be liberal with columns; utils will just append the row.
    # append_submission_xlsx(
    #     xlsx_path,
    #     [
    #         datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    #         name,
    #         email,
    #         e164_phone or phone,
    #         alpha2,
    #         country_name,
    #         dial_code,
    #         service,
    #         message,
    #         request.META.get("REMOTE_ADDR", ""),
    #         request.META.get("HTTP_REFERER", ""),
    #     ],
    # )

    # --- Email notification ---
    subject = f"[Website] Consulting request: {name} – {service or 'General'}"
    text_body = "\n".join([
        "A new consulting request was submitted:",
        f"Name: {name}",
        f"Email: {email}",
        f"Phone: {e164_phone or phone} ({dial_code})",
        f"Country: {country_name or country}",
        f"Service: {service}",
        "",
        "Message:",
        message or "(none)",
        "",
        f"From: {request.META.get('HTTP_REFERER','')}",
        f"IP:   {request.META.get('REMOTE_ADDR','')}",
    ])
    # Reuse your async sender
    _send_contact_email_async(subject, text_body, None)

    messages.success(request, "Thanks! Your request was submitted successfully.")
    return redirect(reverse("cmmsApp:contact_thanks"))

def neplan_electricity(request):
    # Build once here
    countries = []
    for c in pycountry.countries:
        try:
            cc = phonenumbers.country_code_for_region(c.alpha_2)
        except Exception:
            cc = None
        if cc:
            countries.append({"alpha2": c.alpha_2, "name": c.name, "dial": f"+{cc}"})
    countries.sort(key=lambda x: x["name"])

    return render(request, "neplan-electricity.html", {"countries": countries})


def country_list(request):
  """Return [{alpha2,name,dial}] sorted by name."""
  data = []
  for c in pycountry.countries:
      try:
          cc = phonenumbers.country_code_for_region(c.alpha_2)
      except Exception:
          cc = None
      if cc:
          data.append({"alpha2": c.alpha_2, "name": c.name, "dial": f"+{cc}"})
  data.sort(key=lambda x: x["name"])
  return JsonResponse(data, safe=False)
def contact_thanks(request):
    return render(request, "contact_thanks.html", {})
