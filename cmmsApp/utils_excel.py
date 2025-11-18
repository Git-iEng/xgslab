from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

HEADERS = [
    "Timestamp", "First Name", "Last Name", "Company",
    "Email", "Country (alpha-2)", "Country Name",
    "Phone (E.164)", "Message"
]

def append_submission_xlsx(xlsx_path: Path, row: list):
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Submissions"
        ws.append(HEADERS)
        # widen a little
        widths = [20, 18, 16, 24, 28, 18, 28, 22, 40]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        wb.save(xlsx_path)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    ws.append(row)
    wb.save(xlsx_path)
